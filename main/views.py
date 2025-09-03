from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import VardiyaCizelgesi, AracTakip, Arac
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Ana sayfa ve vardiya
def home(request):
    return render(request, 'index.html')

def vardiya(request):
    return render(request, 'vardiya_cizelgesi.html')

# Araç takip sayfası
def arac_takip_view(request):
    araclar = Arac.objects.all()
    return render(request, 'arabalar.html', {'araclar': araclar})

# Vardiya çizelgesi kaydet/load/export
@csrf_exempt
def kaydet_cizelge(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ay = data.get('month')
            yil = data.get('year')
            vardiya_verisi = data.get('data')

            if not all([ay, yil, vardiya_verisi]):
                return JsonResponse({'success': False, 'error': 'Eksik veri.'}, status=400)

            VardiyaCizelgesi.objects.update_or_create(
                ay=ay,
                yil=yil,
                defaults={'veri': vardiya_verisi}
            )
            return JsonResponse({'success': True, 'message': 'Çizelge başarıyla kaydedildi.'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Geçersiz JSON verisi.'}, status=400)
    return JsonResponse({'success': False, 'error': 'Sadece POST metodu desteklenir.'}, status=405)

def yukle_cizelge(request):
    ay = request.GET.get('month')
    yil = request.GET.get('year')
    try:
        cizelge = VardiyaCizelgesi.objects.get(ay=ay, yil=yil)
        return JsonResponse({'success': True, 'data': cizelge.veri})
    except VardiyaCizelgesi.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Bu aya ait çizelge bulunamadı.'}, status=404)

def export_cizelge_excel(request):
    ay = request.GET.get('month')
    yil = request.GET.get('year')

    if not ay or not yil:
        return HttpResponse("Ay ve Yıl parametreleri gereklidir.", status=400)

    vardiya_verisi = []
    try:
        cizelge = VardiyaCizelgesi.objects.get(ay=ay, yil=yil)
        vardiya_verisi = cizelge.veri
    except VardiyaCizelgesi.DoesNotExist:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = f"Vardiya_{ay}_{yil}"

    headers = ['ID', 'Personel']
    if vardiya_verisi:
        num_days = len(vardiya_verisi[0]['shifts'])
        headers += [str(i + 1) for i in range(num_days)]
    else:
        num_days = 0
    ws.append(headers)

    for row_data in vardiya_verisi:
        row = [row_data['id'], row_data['person']] + row_data['shifts']
        ws.append(row)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    for col_idx in range(3, num_days + 3):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=vardiya_cizelgesi_{yil}_{ay}.xlsx'
    wb.save(response)
    return response


def arac_takip_view(request):
    araclar = Arac.objects.filter(aktif=True)
    return render(request, 'arabalar.html', {'araclar': araclar})

@csrf_exempt
def arac_takip_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            plaka = data.get('plaka')
            personel = data.get('personel')
            durum = data.get('durum')
            yorum = data.get('yorum')

            arac_obj = Arac.objects.get(plaka=plaka)

            AracTakip.objects.create(
                arac=arac_obj,
                personel=personel,
                durum=durum,
                yorum=yorum
            )
            return JsonResponse({'success': True})
        except Arac.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Bu plakaya sahip araç bulunamadı.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

def arac_durumu(request):
    plaka = request.GET.get('plaka')
    if not plaka:
        return JsonResponse({'success': False, 'error': 'Plaka parametresi gereklidir.'}, status=400)
    try:
        arac_obj = Arac.objects.get(plaka=plaka)
        son_islem = AracTakip.objects.filter(arac=arac_obj).order_by('-zaman').first()
        if son_islem:
            return JsonResponse({
                'success': True,
                'plaka': son_islem.arac.plaka,
                'personel': son_islem.personel,
                'durum': son_islem.durum,
                'zaman': son_islem.zaman.isoformat(),
                'yorum': son_islem.yorum
            })
        else:
            return JsonResponse({'success': False, 'error': 'Bu araca ait kayıt bulunamadı.'})
    except Arac.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Bu plakaya sahip araç bulunamadı.'}, status=404)